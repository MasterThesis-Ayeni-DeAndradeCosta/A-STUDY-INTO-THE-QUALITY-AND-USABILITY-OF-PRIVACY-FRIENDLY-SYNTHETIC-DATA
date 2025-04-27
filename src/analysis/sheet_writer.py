# file: src/analysis/sheet_writer.py

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

def generate_analysis_sheets(excel_path, df):
    """
    Leaves the 'combined_results' sheet alone.
    Adds or overwrites:
      - Anon_Results (pivot of all anonymized)
      - Synth_Results (pivot of all synthetic)
      - Anon_Best (top 3 per Model & Metric for anonymized)
      - Synth_Best (top 3 per Model & Metric for synthetic)
    """
    wb = openpyxl.load_workbook(excel_path)

    # Remove old versions if they exist
    for sname in ["Anon_Results","Synth_Results","Anon_Best","Synth_Best"]:
        if sname in wb.sheetnames:
            del wb[sname]

    # 1) Full anonymized pivot
    ws_anon = wb.create_sheet("Anon_Results")
    _write_anon_results(ws_anon, df)

    # 2) Full synthetic pivot
    ws_synth = wb.create_sheet("Synth_Results")
    _write_synth_results(ws_synth, df)

    # 3) Full hybrid pivot  --- NEW
    ws_hybrid = wb.create_sheet("Hybrid_Results")
    _write_hybrid_results(ws_hybrid, df)

    # 3) Top 3 anonymized rows per (Model, Metric)
    ws_anon_best = wb.create_sheet("Anon_Best")
    _write_anon_best(ws_anon_best, df)

    # 4) Top 3 synthetic rows per (Model, Metric)
    ws_synth_best = wb.create_sheet("Synth_Best")
    _write_synth_best(ws_synth_best, df)

    # 6) Top 3 hybrid rows --- NEW
    ws_hybrid_best = wb.create_sheet("Hybrid_Best")
    _write_hybrid_best(ws_hybrid_best, df)

    # Auto-size columns
    for sname in ["Anon_Results","Synth_Results","Anon_Best","Synth_Best"]:
        _autosize_worksheet(wb[sname])

    wb.save(excel_path)
    print(f"✅ Created 'Anon_Results','Synth_Results','Anon_Best','Synth_Best' in {excel_path}")


# -------------------------------------------------------------------
# HELPER 1: Full pivot for anonymized data
def _write_anon_results(ws, df):
    sub = df[df["Dataset"]=="Anonymous"]
    if sub.empty:
        ws["A1"] = "No Anonymous rows found."
        return

    metrics = ["Accuracy","Precision","Recall","F1","AUC-ROC"]
    metrics = [m for m in metrics if m in sub.columns]
    if not metrics:
        ws["A1"] = "No metrics found for Anonymous data."
        return

    pivot_anon = (
        sub.groupby(["k_anonymity","l_diversity","suppression_limit","Model"])[metrics]
        .mean(numeric_only=True)
        .reset_index()
    )

    _write_df_to_sheet(ws, pivot_anon)


# -------------------------------------------------------------------
# HELPER 2: Full pivot for synthetic data
def _write_synth_results(ws, df):
    """
    Write pivot table summarizing synthetic results (CTGAN, TVAE, GaussianCopula).
    """
    sub = df[df["Dataset"].isin(["CTGAN", "TVAE", "GaussianCopula"])]  # 🔵 Only real synthetics
    if sub.empty:
        ws["A1"] = "No Synthetic rows found."
        return

    keep_cols = ["Dataset", "Model", "epochs", "rows_generated_at_runtime", "Accuracy", "Precision", "Recall", "F1", "AUC-ROC"]
    keep_cols = [c for c in keep_cols if c in sub.columns]

    pivot_synth = sub[keep_cols]

    _write_df_to_sheet(ws, pivot_synth)



# -------------------------------------------------------------------
# HELPER 3: Top 3 anonymized rows per (Model, Metric)
def _write_anon_best(ws, df):
    sub = df[df["Dataset"]=="Anonymous"]
    if sub.empty:
        ws["A1"] = "No Anonymous rows found."
        return

    metrics = ["Accuracy","Precision","Recall","F1","AUC-ROC"]
    metrics = [m for m in metrics if m in sub.columns]

    best_rows = []
    # For each model, for each metric, we gather the top 3 rows
    for model in sub["Model"].unique():
        model_df = sub[sub["Model"]==model]
        for metric in metrics:
            valid_df = model_df[model_df[metric].notna()]
            if valid_df.empty:
                continue
            top_3 = valid_df.nlargest(3, metric)  # sort descending by that metric
            for rank, (idx, row_data) in enumerate(top_3.iterrows(), start=1):
                row_data = row_data.copy()
                row_data["WhichMetric"] = metric
                row_data["Rank"] = rank
                row_data["MetricValue"] = row_data[metric]
                best_rows.append(row_data)

    if not best_rows:
        ws["A1"] = "No best anonymized rows found."
        return

    best_df = pd.DataFrame(best_rows)
    # Reorder columns. You can omit or add more if you want
    keep_cols = ["Model","WhichMetric","Rank","MetricValue",
                 "k_anonymity","l_diversity","suppression_limit",
                 "Accuracy","Precision","Recall","F1","AUC-ROC"]
    keep_cols = [c for c in keep_cols if c in best_df.columns]
    best_df = best_df[keep_cols]

    # Sort so it’s grouped by Model, then Metric, then Rank
    best_df = best_df.sort_values(by=["Model","WhichMetric","Rank"], ascending=True)

    _write_df_to_sheet(ws, best_df)


# -------------------------------------------------------------------
# HELPER 4: Top 3 synthetic rows per (Model, Metric)
def _write_synth_best(ws, df):
    """
    Write pivot table summarizing the best synthetic results per dataset/model.
    """
    sub = df[df["Dataset"].isin(["CTGAN", "TVAE", "GaussianCopula"])]  # 🔵 Only real synthetics
    if sub.empty:
        ws["A1"] = "No Synthetic rows found."
        return

    metrics = ["Accuracy", "Precision", "Recall", "F1", "AUC-ROC"]
    metrics = [m for m in metrics if m in sub.columns]

    best_rows = []
    for model in sub["Model"].unique():
        model_df = sub[sub["Model"] == model]
        for metric in metrics:
            valid_df = model_df[model_df[metric].notna()]
            if valid_df.empty:
                continue
            top_3 = valid_df.nlargest(3, metric)
            for rank, (idx, row_data) in enumerate(top_3.iterrows(), start=1):
                row_data = row_data.copy()
                row_data["WhichMetric"] = metric
                row_data["Rank"] = rank
                row_data["MetricValue"] = row_data[metric]
                best_rows.append(row_data)

    if not best_rows:
        ws["A1"] = "No best synthetic rows found."
        return

    best_df = pd.DataFrame(best_rows)

    keep_cols = ["Dataset", "Model", "WhichMetric", "Rank", "MetricValue",
                 "epochs", "rows_generated_at_runtime",
                 "Accuracy", "Precision", "Recall", "F1", "AUC-ROC"]
    keep_cols = [c for c in keep_cols if c in best_df.columns]

    best_df = best_df[keep_cols]

    best_df = best_df.sort_values(by=["Model", "WhichMetric", "Rank"], ascending=True)

    _write_df_to_sheet(ws, best_df)


def _write_hybrid_results(ws, df):
    """
    Write pivot table summarizing Hybrid results (e.g., CTGAN_HYBRID, TVAE_HYBRID).
    """
    sub = df[df["Dataset"].str.endswith("_HYBRID")]
    if sub.empty:
        ws["A1"] = "No Hybrid rows found."
        return

    keep_cols = ["Dataset", "Model", "k_anonymity", "l_diversity", "suppression_limit",
                 "epochs", "rows_generated_at_runtime", "Accuracy", "Precision", "Recall", "F1", "AUC-ROC"]

    keep_cols = [col for col in keep_cols if col in sub.columns]

    pivot_hybrid = sub[keep_cols]

    _write_df_to_sheet(ws, pivot_hybrid)


def _write_hybrid_best(ws, df):
    """
    Write pivot table summarizing the best Hybrid results per dataset/model.
    """
    sub = df[df["Dataset"].str.endswith("_HYBRID")]
    if sub.empty:
        ws["A1"] = "No Hybrid rows found."
        return

    metrics = ["Accuracy", "Precision", "Recall", "F1", "AUC-ROC"]
    metrics = [m for m in metrics if m in sub.columns]

    best_rows = []
    for model in sub["Model"].unique():
        model_df = sub[sub["Model"] == model]
        for metric in metrics:
            valid_df = model_df[model_df[metric].notna()]
            if valid_df.empty:
                continue
            top_3 = valid_df.nlargest(3, metric)
            for rank, (idx, row_data) in enumerate(top_3.iterrows(), start=1):
                row_data = row_data.copy()
                row_data["WhichMetric"] = metric
                row_data["Rank"] = rank
                row_data["MetricValue"] = row_data[metric]
                best_rows.append(row_data)

    if not best_rows:
        ws["A1"] = "No best Hybrid rows found."
        return

    best_df = pd.DataFrame(best_rows)

    keep_cols = ["Dataset", "Model", "WhichMetric", "Rank", "MetricValue",
                 "k_anonymity", "l_diversity", "suppression_limit", "epochs", "rows_generated_at_runtime",
                 "Accuracy", "Precision", "Recall", "F1", "AUC-ROC"]

    keep_cols = [col for col in keep_cols if col in best_df.columns]

    best_df = best_df[keep_cols]

    best_df = best_df.sort_values(by=["Model", "WhichMetric", "Rank"], ascending=True)

    _write_df_to_sheet(ws, best_df)



# -------------------------------------------------------------------
# Writes a DataFrame to an openpyxl worksheet (row=1 col=1)
def _write_df_to_sheet(ws, df, start_row=1, start_col=1):
    for c_idx, col_name in enumerate(df.columns, start_col):
        ws.cell(row=start_row, column=c_idx, value=col_name)
    for r_idx, row_data in enumerate(df.itertuples(index=False), start_row+1):
        for c_idx, val in enumerate(row_data, start_col):
            ws.cell(row=r_idx, column=c_idx, value=val)


def _autosize_worksheet(ws):
    for col in range(1, ws.max_column+1):
        max_length = 0
        col_letter = get_column_letter(col)
        for row in range(1, ws.max_row+1):
            cell_val = ws.cell(row=row, column=col).value
            if cell_val is not None:
                length = len(str(cell_val))
                if length > max_length:
                    max_length = length
        ws.column_dimensions[col_letter].width = max_length + 2
